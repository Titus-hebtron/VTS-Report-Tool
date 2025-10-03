
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import Workbook
from io import BytesIO

def extract_incident_info(df):
    info = {}
    for i, row in df.iterrows():
        for col in row.index:
            val = str(row[col])
            if 'Incident Notification Date' in val:
                info['Incident Date'] = row[col+1] if col+1 < len(row) else ''
            if 'Incident Notification Time' in val:
                info['Incident Time'] = row[col+1] if col+1 < len(row) else ''
            if 'Location of Incident' in val:
                info['Location'] = row[col+1] if col+1 < len(row) else ''
            if 'Type of Vehicle' in val:
                info['Vehicle Type'] = row[col+1] if col+1 < len(row) else ''
            if 'Number of Accident Vehicles' in val:
                info['Number of Vehicles'] = row[col+1] if col+1 < len(row) else ''
            if 'Number of Injured People' in val:
                info['Number Injured'] = row[col+1] if col+1 < len(row) else ''
            if 'Conditions of Accident Vehicles' in val:
                info['Vehicle Condition'] = row[col+1] if col+1 < len(row) else ''
            if 'Clearing Time' in val:
                info['Clearing Time'] = row[col+1] if col+1 < len(row) else ''
            if 'Caller' in val:
                info['Caller'] = row[col+1] if col+1 < len(row) else ''
            if 'Phone Number' in val:
                info['Phone Number'] = row[col+1] if col+1 < len(row) else ''
            if 'Nature of Incident' in val:
                info['Nature of Incident'] = row[col+1] if col+1 < len(row) else ''
            if 'Bound:' in val:
                info['Bound'] = row[col+1] if col+1 < len(row) else ''
            if 'Chainage:' in val:
                info['Chainage'] = row[col+1] if col+1 < len(row) else ''
            if 'Response Time' in val:
                info['Response Time'] = row[col+1] if col+1 < len(row) else ''
    return info

def create_excel_report(info):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Incident Report'
    from openpyxl.styles import Font, Alignment, PatternFill
    yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    # Row 7: Incident Notification Date, Incident Notification Time
    ws['B7'] = 'Incident Notification Date:'
    ws['C7'] = info.get('Incident Date', '')
    ws['F7'] = 'Incident Notification Time:'
    ws['G7'] = info.get('Incident Time', '')
    ws['C7'].fill = yellow
    ws['G7'].fill = yellow
    # Row 8: Incident Information Resource (checkboxes), Caller, Phone Number
    ws['B8'] = 'Incident Information Resource:'
    ws['C8'] = '✔' if info.get('KeNHA', False) else ''
    ws['D8'] = '✔' if info.get('Police', False) else ''
    ws['E8'] = '✔' if info.get('Engineer', False) else ''
    ws['F8'] = '✔' if info.get('Road Users', False) else ''
    ws['G8'] = '✔' if info.get('Highway Patrol Unit', False) else ''
    ws['H8'] = '✔' if info.get('Others', False) else ''
    ws['I8'] = 'Caller:'
    ws['J8'] = info.get('Caller', '')
    ws['J8'].fill = yellow
    ws['L8'] = 'Phone Number:'
    ws['M8'] = info.get('Phone Number', '')
    ws['M8'].fill = yellow
    # Row 9: Nature of Incident (checkboxes)
    ws['B9'] = 'Nature of Incident:'
    ws['C9'] = '✔' if info.get('Accident', False) else ''
    ws['D9'] = 'Accident'
    ws['E9'] = '✔' if info.get('Emergency', False) else ''
    ws['F9'] = 'Emergency'
    ws['G9'] = '✔' if info.get('Others (Nature)', False) else ''
    ws['H9'] = 'Others'
    # Row 10: Location, Bound, Chainage
    ws['B10'] = 'Location of Incident :'
    ws['C10'] = info.get('Location', '')
    ws['C10'].fill = yellow
    ws['E10'] = 'Bound:'
    ws['F10'] = info.get('Bound', '')
    ws['F10'].fill = yellow
    ws['H10'] = 'Chainage:'
    ws['I10'] = info.get('Chainage', '')
    ws['I10'].fill = yellow
    # Row 11: Number of Accident Vehicles, Type of Vehicle
    ws['B11'] = 'Number of Accident Vehicles:'
    ws['C11'] = info.get('Number of Vehicles', '')
    ws['C11'].fill = yellow
    ws['E11'] = 'Type of Vehicle:'
    ws['F11'] = info.get('Vehicle Type', '')
    ws['F11'].fill = yellow
    # Row 12: Conditions of Accident Vehicles, Number of Injured People
    ws['B12'] = 'Conditions of Accident Vehicles:'
    ws['C12'] = info.get('Vehicle Condition', '')
    ws['C12'].fill = yellow
    ws['H12'] = 'Number of Injured People:'
    ws['I12'] = info.get('Number Injured', '')
    ws['I12'].fill = yellow
    # Row 13: Conditions of Injured People, The Injured Part
    ws['B13'] = 'Conditions of Injured People:'
    ws['C13'] = info.get('Conditions of Injured People', '')
    ws['C13'].fill = yellow
    ws['F13'] = 'The Injured Part:'
    ws['G13'] = info.get('The Injured Part', '')
    ws['G13'].fill = yellow
    # Row 14: Fire Hazard, Oil Leakage, Chemical Leakage
    ws['B14'] = 'Fire Hazard:'
    ws['C14'] = info.get('Fire Hazard', '')
    ws['C14'].fill = yellow
    ws['E14'] = 'Oil Leakage:'
    ws['F14'] = info.get('Oil Leakage', '')
    ws['F14'].fill = yellow
    ws['H14'] = 'Chemical Leakage:'
    ws['I14'] = info.get('Chemical Leakage', '')
    ws['I14'].fill = yellow
    # Row 15: Damage To Road Furniture
    ws['B15'] = 'Damage To Road Furniture:'
    ws['C15'] = info.get('Damage To Road Furniture', '')
    ws['C15'].fill = yellow
    # Row 16: Response Time, Clearing Time
    ws['B16'] = 'Response Time:'
    ws['C16'] = info.get('Response Time', '')
    ws['C16'].fill = yellow
    ws['F16'] = 'Clearing Time:'
    ws['G16'] = info.get('Clearing Time', '')
    ws['G16'].fill = yellow
    # Row 17: Department Contact (checkboxes)
    ws['B17'] = 'Department Contact'
    ws['C17'] = '✔' if info.get('KeNHA (Dept)', False) else ''
    ws['D17'] = '✔' if info.get('Police (Dept)', False) else ''
    ws['E17'] = '✔' if info.get('Engineer (Dept)', False) else ''
    ws['F17'] = '✔' if info.get('Highway Patrol Unit (Dept)', False) else ''
    ws['G17'] = '✔' if info.get('Ambulance', False) else ''
    ws['H17'] = '✔' if info.get('Recovery Vehicle', False) else ''
    # Formatting: bold for all labels
    for row in ws.iter_rows(min_row=7, max_row=17, min_col=2, max_col=14):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and (cell.value.endswith(':') or cell.value in ['Accident', 'Emergency', 'Others']):
                cell.font = Font(bold=True)
    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    # Insert images if present
    if 'Images' in info and info['Images']:
        from openpyxl.drawing.image import Image as XLImage
        img_row = 19
        img_col = 2  # Default starting column B
        for i, img in enumerate(info['Images']):
            if img is not None:
                try:
                    img.seek(0)
                    img_bytes = img.read()
                    import tempfile
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_img:
                        tmp_img.write(img_bytes)
                        tmp_img.flush()
                        xl_img = XLImage(tmp_img.name)
                        xl_img.width = 180  # px (enlarged)
                        xl_img.height = 135  # px (enlarged)
                        if i == 1:  # Second image in column C
                            img_col = 3
                        elif i == 2:  # Third image in column E
                            img_col = 5
                        cell = ws.cell(row=img_row, column=img_col)
                        ws.add_image(xl_img, cell.coordinate)
                        img_col += 2  # Next image to the right (skip a column for spacing)
                except Exception as e:
                    pass
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

st.title('Online Fillable Incident Report')

uploaded_file = st.file_uploader('Upload Incident Report CSV (optional)', type=['csv'])
info = {}
if uploaded_file:
    df = pd.read_csv(uploaded_file, header=None)
    info = extract_incident_info(df)


with st.form('incident_form'):
    st.write('Fill in the incident report fields:')
    # Row 1: Incident Notification Date, Incident Notification Time
    c1, c2 = st.columns(2)
    with c1:
        incident_date = st.text_input('Incident Notification Date', value=info.get('Incident Date', ''))
    with c2:
        incident_time = st.text_input('Incident Notification Time', value=info.get('Incident Time', ''))

    # Row 2: Incident Information Resource (checkboxes)
    st.markdown('**Incident Information Resource:**')
    c3, c4, c5, c6, c7, c8 = st.columns(6)
    with c3:
        res_kenha = st.checkbox('KeNHA', value=info.get('KeNHA', False))
    with c4:
        res_police = st.checkbox('Police', value=info.get('Police', False))
    with c5:
        res_engineer = st.checkbox('Engineer', value=info.get('Engineer', False))
    with c6:
        res_road_users = st.checkbox('Road Users', value=info.get('Road Users', False))
    with c7:
        res_highway_patrol = st.checkbox('Highway Patrol Unit', value=info.get('Highway Patrol Unit', False))
    with c8:
        res_others = st.checkbox('Others', value=info.get('Others', False))

    # Row 3: Caller, Phone Number
    c9, c10 = st.columns(2)
    with c9:
        caller = st.text_input('Caller', value=info.get('Caller', ''))
    with c10:
        phone_number = st.text_input('Phone Number', value=info.get('Phone Number', ''))

    # Row 4: Nature of Incident (checkboxes)
    st.markdown('**Nature of Incident:**')
    c11, c12, c13 = st.columns(3)
    with c11:
        nature_accident = st.checkbox('Accident', value=info.get('Accident', False))
    with c12:
        nature_emergency = st.checkbox('Emergency', value=info.get('Emergency', False))
    with c13:
        nature_others = st.checkbox('Others (Nature)', value=info.get('Others (Nature)', False))

    # Row 5: Location, Bound, Chainage
    c14, c15, c16 = st.columns(3)
    with c14:
        location = st.text_input('Location of Incident', value=info.get('Location', ''))
    with c15:
        bound = st.text_input('Bound', value=info.get('Bound', ''))
    with c16:
        chainage = st.text_input('Chainage', value=info.get('Chainage', ''))

    # Row 6: Number of Accident Vehicles, Type of Vehicle
    c17, c18 = st.columns(2)
    with c17:
        num_vehicles = st.text_input('Number of Accident Vehicles', value=info.get('Number of Vehicles', ''))
    with c18:
        vehicle_type = st.text_input('Type of Vehicle', value=info.get('Vehicle Type', ''))

    # Row 7: Conditions of Accident Vehicles, Number of Injured People
    c19, c20 = st.columns(2)
    with c19:
        vehicle_condition = st.text_input('Conditions of Accident Vehicles', value=info.get('Vehicle Condition', ''))
    with c20:
        num_injured = st.text_input('Number of Injured People', value=info.get('Number Injured', ''))

    # Row 8: Conditions of Injured People, The Injured Part
    c21, c22 = st.columns(2)
    with c21:
        cond_injured = st.text_input('Conditions of Injured People', value=info.get('Conditions of Injured People', ''))
    with c22:
        injured_part = st.text_input('The Injured Part', value=info.get('The Injured Part', ''))

    # Row 9: Fire Hazard, Oil Leakage, Chemical Leakage
    c23, c24, c25 = st.columns(3)
    with c23:
        fire_hazard = st.text_input('Fire Hazard', value=info.get('Fire Hazard', ''))
    with c24:
        oil_leakage = st.text_input('Oil Leakage', value=info.get('Oil Leakage', ''))
    with c25:
        chemical_leakage = st.text_input('Chemical Leakage', value=info.get('Chemical Leakage', ''))

    # Row 10: Damage To Road Furniture
    damage_road_furniture = st.text_input('Damage To Road Furniture', value=info.get('Damage To Road Furniture', ''))

    # Row 11: Response Time, Clearing Time
    c26, c27 = st.columns(2)
    with c26:
        response_time = st.text_input('Response Time', value=info.get('Response Time', ''))
    with c27:
        clearing_time = st.text_input('Clearing Time', value=info.get('Clearing Time', ''))

    # Row 12: Department Contact (checkboxes)
    st.markdown('**Department Contact:**')
    c28, c29, c30, c31, c32, c33 = st.columns(6)
    with c28:
        dep_kenha = st.checkbox('KeNHA (Dept)', value=info.get('KeNHA (Dept)', False))
    with c29:
        dep_police = st.checkbox('Police (Dept)', value=info.get('Police (Dept)', False))
    with c30:
        dep_engineer = st.checkbox('Engineer (Dept)', value=info.get('Engineer (Dept)', False))
    with c31:
        dep_highway_patrol = st.checkbox('Highway Patrol Unit (Dept)', value=info.get('Highway Patrol Unit (Dept)', False))
    with c32:
        dep_ambulance = st.checkbox('Ambulance', value=info.get('Ambulance', False))
    with c33:
        dep_recovery = st.checkbox('Recovery Vehicle', value=info.get('Recovery Vehicle', False))

    image_files = st.file_uploader('Upload Vehicle Images (multiple allowed)', type=['jpg', 'jpeg', 'png'], accept_multiple_files=True, help='You can take a photo or select from your device.')
    submitted = st.form_submit_button('Generate Excel Report')

if 'submitted' in locals() and submitted:
    # Minimal test PDF download (for debugging Streamlit download)
    import base64
    minimal_pdf_bytes = b'%PDF-1.4\n1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 200 200] /Contents 4 0 R >>\nendobj\n4 0 obj\n<< /Length 44 >>\nstream\nBT /F1 24 Tf 72 120 Td (Hello PDF) Tj ET\nendstream\nendobj\nxref\n0 5\n0000000000 65535 f \n0000000010 00000 n \n0000000060 00000 n \n0000000117 00000 n \n0000000212 00000 n \ntrailer\n<< /Size 5 /Root 1 0 R >>\nstartxref\n312\n%%EOF'
    st.download_button(
        label='Download Minimal Test PDF',
        data=minimal_pdf_bytes,
        file_name='test_minimal.pdf',
        mime='application/pdf'
    )
    filled_info = {
        'Incident Date': incident_date,
        'Incident Time': incident_time,
        'KeNHA': res_kenha,
        'Police': res_police,
        'Engineer': res_engineer,
        'Road Users': res_road_users,
        'Highway Patrol Unit': res_highway_patrol,
        'Others': res_others,
        'Caller': caller,
        'Phone Number': phone_number,
        'Accident': nature_accident,
        'Emergency': nature_emergency,
        'Others (Nature)': nature_others,
        'Nature of Incident': '',
        'Location': location,
        'Bound': bound,
        'Chainage': chainage,
        'Number of Vehicles': num_vehicles,
        'Vehicle Type': vehicle_type,
        'Vehicle Condition': vehicle_condition,
        'Number Injured': num_injured,
        'Conditions of Injured People': cond_injured,
        'The Injured Part': injured_part,
        'Fire Hazard': fire_hazard,
        'Oil Leakage': oil_leakage,
        'Chemical Leakage': chemical_leakage,
        'Damage To Road Furniture': damage_road_furniture,
        'Response Time': response_time,
        'Clearing Time': clearing_time,
        'KeNHA (Dept)': dep_kenha,
        'Police (Dept)': dep_police,
        'Engineer (Dept)': dep_engineer,
        'Highway Patrol Unit (Dept)': dep_highway_patrol,
        'Ambulance': dep_ambulance,
        'Recovery Vehicle': dep_recovery,
        'Images': image_files
    }
    excel_data = create_excel_report(filled_info)
    excel_data.seek(0)
    st.success('Incident report ready!')
    st.download_button(
        label='Download Completed Incident Report Excel',
        data=excel_data,
        file_name='incident_report_filled.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # PDF generation
    from fpdf import FPDF
    import tempfile
    import os
    import atexit
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    font_path = os.path.join(os.path.dirname(__file__), 'DejaVuSans.ttf')
    font_bold_path = os.path.join(os.path.dirname(__file__), 'DejaVuSans-Bold.ttf')
    if not os.path.exists(font_bold_path):
        st.error('DejaVuSans-Bold.ttf font file is missing. Please download it from https://dejavu-fonts.github.io/ and place it in your project directory.')
        st.stop()
    if not os.path.exists(font_path):
        st.error('DejaVuSans.ttf font file is missing. Please download it from https://dejavu-fonts.github.io/ and place it in your project directory.')
        st.stop()
    pdf.add_font('DejaVu', '', font_path, uni=True)
    pdf.add_font('DejaVu', 'B', font_bold_path, uni=True)
    pdf.set_font('DejaVu', '', 14)
    pdf.add_page()
    left_margin = 10
    top_margin = 10
    page_width = pdf.w - 2 * left_margin
    page_height = pdf.h - 2 * top_margin
    pdf.rect(left_margin, top_margin, page_width, page_height)
    pdf.set_xy(left_margin, top_margin)
    pdf.cell(page_width, 10, 'Accident/Incident Report', ln=1, align='C')
    pdf.set_font('DejaVu', '', 10)
    y = pdf.get_y() + 2
    pdf.set_y(y)
    def pdf_checkbox(val):
        return '[✔]' if val else '[ ]'
    x0 = left_margin
    y0 = pdf.get_y()
    row_h = 8
    pdf.set_xy(x0, y0)
    pdf.cell(60, row_h, 'Incident Notification Date:', 1)
    pdf.cell(40, row_h, filled_info['Incident Date'], 1)
    pdf.cell(60, row_h, 'Incident Notification Time:', 1)
    pdf.cell(40, row_h, filled_info['Incident Time'], 1)
    pdf.ln(row_h)
    pdf.set_x(x0)
    pdf.cell(60, row_h, 'Incident Information Resource:', 1)
    pdf.cell(25, row_h, 'KeNHA ' + pdf_checkbox(filled_info['KeNHA']), 1)
    pdf.cell(25, row_h, 'Police ' + pdf_checkbox(filled_info['Police']), 1)
    pdf.cell(30, row_h, 'Engineer ' + pdf_checkbox(filled_info['Engineer']), 1)
    pdf.cell(35, row_h, 'Road Users ' + pdf_checkbox(filled_info['Road Users']), 1)
    pdf.cell(40, row_h, 'Highway Patrol ' + pdf_checkbox(filled_info['Highway Patrol Unit']), 1)
    pdf.cell(30, row_h, 'Others ' + pdf_checkbox(filled_info['Others']), 1)
    pdf.ln(row_h)
    pdf.set_x(x0)
    pdf.cell(20, row_h, 'Caller:', 1)
    pdf.cell(60, row_h, filled_info['Caller'], 1)
    pdf.cell(30, row_h, 'Phone Number:', 1)
    pdf.cell(60, row_h, filled_info['Phone Number'], 1)
    pdf.ln(row_h)
    pdf.set_x(x0)
    pdf.cell(40, row_h, 'Nature of Incident:', 1)
    pdf.cell(40, row_h, 'Accident ' + pdf_checkbox(filled_info['Accident']), 1)
    pdf.cell(40, row_h, 'Emergency ' + pdf_checkbox(filled_info['Emergency']), 1)
    pdf.cell(40, row_h, 'Others ' + pdf_checkbox(filled_info['Others (Nature)']), 1)
    pdf.ln(row_h)
    pdf.set_x(x0)
    pdf.cell(40, row_h, 'Location:', 1)
    pdf.cell(60, row_h, filled_info['Location'], 1)
    pdf.cell(25, row_h, 'Bound:', 1)
    pdf.cell(30, row_h, filled_info['Bound'], 1)
    pdf.cell(25, row_h, 'Chainage:', 1)
    pdf.cell(30, row_h, filled_info['Chainage'], 1)
    pdf.ln(row_h)
    pdf.set_x(x0)
    pdf.cell(60, row_h, 'Number of Accident Vehicles:', 1)
    pdf.cell(40, row_h, filled_info['Number of Vehicles'], 1)
    pdf.cell(40, row_h, 'Type of Vehicle:', 1)
    pdf.cell(60, row_h, filled_info['Vehicle Type'], 1)
    pdf.ln(row_h)
    pdf.set_x(x0)
    pdf.cell(60, row_h, 'Conditions of Accident Vehicles:', 1)
    pdf.cell(60, row_h, filled_info['Vehicle Condition'], 1)
    pdf.cell(50, row_h, 'Number of Injured People:', 1)
    pdf.cell(30, row_h, filled_info['Number Injured'], 1)
    pdf.ln(row_h)
    pdf.set_x(x0)
    pdf.cell(60, row_h, 'Conditions of Injured People:', 1)
    pdf.cell(60, row_h, filled_info['Conditions of Injured People'], 1)
    pdf.cell(40, row_h, 'The Injured Part:', 1)
    pdf.cell(40, row_h, filled_info['The Injured Part'], 1)
    pdf.ln(row_h)
    pdf.set_x(x0)
    pdf.cell(40, row_h, 'Fire Hazard:', 1)
    pdf.cell(40, row_h, filled_info['Fire Hazard'], 1)
    pdf.cell(40, row_h, 'Oil Leakage:', 1)
    pdf.cell(40, row_h, filled_info['Oil Leakage'], 1)
    pdf.cell(40, row_h, 'Chemical Leakage:', 1)
    pdf.cell(40, row_h, filled_info['Chemical Leakage'], 1)
    pdf.ln(row_h)
    pdf.set_x(x0)
    pdf.cell(60, row_h, 'Damage To Road Furniture:', 1)
    pdf.cell(120, row_h, filled_info['Damage To Road Furniture'], 1)
    pdf.ln(row_h)
    pdf.set_x(x0)
    pdf.cell(40, row_h, 'Response Time:', 1)
    pdf.cell(40, row_h, filled_info['Response Time'], 1)
    pdf.cell(40, row_h, 'Clearing Time:', 1)
    pdf.cell(40, row_h, filled_info['Clearing Time'], 1)
    pdf.ln(row_h)
    pdf.set_x(x0)
    pdf.cell(60, row_h, 'Department Contact:', 1)
    pdf.cell(28, row_h, 'KeNHA ' + pdf_checkbox(filled_info['KeNHA (Dept)']), 1)
    pdf.cell(28, row_h, 'Police ' + pdf_checkbox(filled_info['Police (Dept)']), 1)
    pdf.cell(32, row_h, 'Engineer ' + pdf_checkbox(filled_info['Engineer (Dept)']), 1)
    pdf.cell(38, row_h, 'Highway Patrol ' + pdf_checkbox(filled_info['Highway Patrol Unit (Dept)']), 1)
    pdf.cell(32, row_h, 'Ambulance ' + pdf_checkbox(filled_info['Ambulance']), 1)
    pdf.cell(40, row_h, 'Recovery Vehicle ' + pdf_checkbox(filled_info['Recovery Vehicle']), 1)
    pdf.ln(row_h)
    temp_image_paths = []
    if filled_info.get('Images'):
        pdf.ln(3)
        pdf.set_font('DejaVu', 'B', 11)
        pdf.cell(0, 8, 'Incident Images:', ln=1)
        images = [img for img in filled_info['Images'] if img is not None]
        if images:
            max_width = page_width - 4  # small padding
            gap = 6  # gap between images in mm
            max_img_w = 50  # max width per image
            x_img = left_margin + 2
            y_img = pdf.get_y() + 2
            row_height = 0
            images_in_row = []
            for img in images:
                try:
                    img.seek(0)
                    img_bytes = img.read()
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_img:
                        tmp_img.write(img_bytes)
                        tmp_img.flush()
                        temp_image_paths.append(tmp_img.name)
                        # Check if next image fits in current row
                        if (x_img + max_img_w) > (left_margin + page_width):
                            # Move to next row
                            x_img = left_margin + 2
                            y_img += row_height + gap
                            row_height = 0
                        pdf.image(tmp_img.name, x=x_img, y=y_img, w=max_img_w)
                        img_h = max_img_w * 0.75  # estimate height
                        row_height = max(row_height, img_h)
                        x_img += max_img_w + gap
                except Exception as e:
                    st.warning(f"Could not add image to PDF: {e}")
            pdf.ln(row_height + 6)
    def cleanup_temp_images():
        for path in temp_image_paths:
            try:
                os.remove(path)
            except Exception:
                pass
    atexit.register(cleanup_temp_images)
    try:
        pdf_raw = pdf.output(dest='S')
        if isinstance(pdf_raw, str):
            pdf_data = pdf_raw.encode('latin1')
        else:
            pdf_data = bytes(pdf_raw)
        pdf_size = len(pdf_data)
        st.info(f"PDF size: {pdf_size} bytes. First 10 bytes: {pdf_data[:10]}")
        if pdf_size == 0:
            st.error("PDF generation failed: The file is empty. Please check for missing fonts or data.")
        else:
            st.download_button(
                label='Download Incident Report PDF',
                data=pdf_data,
                file_name='incident_report.pdf',
                mime='application/pdf'
            )
    except Exception as e:
        st.error(f"PDF generation failed: {e}")
