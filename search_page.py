def fill_incident_template(ws, row, index):
    from openpyxl.styles import Alignment, Border, Side, Font
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl import utils

    # Define merged cells
    ws.merge_cells('A1:K1')
    ws.merge_cells('B2:K2')
    ws.merge_cells('A3:K3')
    ws.merge_cells('J4:K4')
    ws.merge_cells('B9:C9')
    ws.merge_cells('F9:H9')
    ws.merge_cells('C10:K10')
    ws.merge_cells('C11:D11')
    ws.merge_cells('C13:K13')
    ws.merge_cells('D15:K15')
    ws.merge_cells('D16:K16')

    # Determine contractor
    contractor_id = row.get('contractor_id')
    if contractor_id == 1:
        contractor = 'wizpro'
        rfi_prefix = 'WIZP'
    elif contractor_id == 2:
        contractor = 'paschal'
        rfi_prefix = 'PASC'
    else:
        contractor = 'paschal'
        rfi_prefix = 'PASC'
    CONTRACTOR_NAME = contractor

    # Logo insertion
    ORG_LOGO_PATH = 'Kenhalogo.png'
    PASCHAL_LOGO_PATH = 'paschal_logo.png'  # Placeholder, ensure file exists
    WIZPRO_LOGO_PATH = 'wizpro_logo.png'    # Placeholder

    # Insert Main Organization Logo
    try:
        img = OpenpyxlImage(ORG_LOGO_PATH)
        img.width = 1000  # Increase width for visibility
        img.height = 100  # Increase height for visibility
        img.anchor = 'B1'
        ws.add_image(img)
        ws.row_dimensions[1].height = 100  # Set row height to fit larger logo
    except Exception as e:
        print(f"Warning: Could not insert org logo. {e}")

    # Insert Contractor Logo
    try:
        if CONTRACTOR_NAME.lower() == 'paschal':
            img = OpenpyxlImage(PASCHAL_LOGO_PATH)
            img.width = 1000
            img.height = 100
            img.anchor = 'B2'
            ws.add_image(img)
            ws.row_dimensions[2].height = 60  # Set row height to fit logo
        elif CONTRACTOR_NAME.lower() == 'wizpro':
            img = OpenpyxlImage(WIZPRO_LOGO_PATH)
            img.width = 1000
            img.height = 100
            img.anchor = 'B2'
            ws.add_image(img)
            ws.row_dimensions[2].height = 60
    except Exception as e:
        print(f"Warning: Could not insert contractor logo. {e}")

    # Set title based on contractor
    if contractor == 'paschal':
        title = 'PERFORMANCE BASED MAINTENANCE CONTRACT FOR RUIRU-THIKA HIGHWAY'
    elif contractor == 'wizpro':
        title = 'PERFORMANCE BASED MAINTENANCE CONTRACT FOR NAIROBI-RUIRU'
    else:
        title = 'INCIDENT REPORT'
    ws['A3'] = title.upper()
    from openpyxl.styles import Alignment
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A3'].font = Font(bold=True, size=12)

    # Styles
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    for cell_range in ['C10', 'C13', 'D15', 'D16']:
        ws[cell_range].alignment = wrap_alignment

    # Data population
    rfi_number = f"{rfi_prefix}{index+1:03d}"
    ws['J4'] = rfi_number

    # Row 6: Date and Time
    ws['A6'] = 'Incident Date:'
    ws['B6'] = row.get('incident_date')
    ws['D6'] = 'Incident Time:'
    ws['E6'] = row.get('incident_time')

    # Row 7: Incident Information Resource
    ws['A7'] = 'Incident Information Resource:'
    ws['B7'] = row.get('incident_info_resource', '')

    # Row 9: Caller Info and Patrol Vehicle
    ws['A9'] = 'Caller:'
    ws['B9'] = row.get('caller')
    ws['E9'] = 'Phone Number:'
    ws['F9'] = row.get('phone_number')
    ws['I9'] = 'Patrol Car:'
    ws['J9'] = row.get('patrol_car')

    # Row 10: Nature of Incident
    ws['A10'] = 'Nature of Incident:'
    nature_text = f"{row.get('incident_type', 'N/A')} - {row.get('description', 'No details provided.')}"
    ws['C10'] = nature_text

    # Row 11: Location Details
    ws['A11'] = 'Location:'
    ws['C11'] = row.get('location')
    ws['E11'] = 'Bound:'
    ws['F11'] = row.get('bound')
    ws['H11'] = 'Chainage:'
    ws['I11'] = row.get('chainage')

    # Row 13: Vehicle Details
    ws['A13'] = 'Vehicle Details:'
    vehicle_details = (
        f"Type: {row.get('vehicle_type', 'N/A')} "
        f"({row.get('num_vehicles', '1')} unit(s)). "
        f"Condition: {row.get('vehicle_condition', 'Unknown')}"
    )
    ws['C13'] = vehicle_details

    # Row 14: Hazards
    ws['A14'] = 'Fire Hazard:'
    ws['C14'] = row.get('fire_hazard', 'No')
    ws['E14'] = 'Oil Leakage:'
    ws['F14'] = row.get('oil_leakage', 'No')
    ws['H14'] = 'Chemical Leakage:'
    ws['I14'] = row.get('chemical_leakage', 'No')

    # Row 15: Injured people
    ws['A15'] = 'Injured People:'
    injured_details = (
        f"No: {row.get('num_injured', '0')}. "
        f"Condition: {row.get('cond_injured', 'N/A')}. "
        f"Part: {row.get('injured_part', 'N/A')}"
    )
    ws['C15'] = injured_details

    # Row 16: Road Furniture Damage
    ws['A16'] = 'Road Furniture Damage:'
    ws['C16'] = row.get('damage_road_furniture', 'Nil')

    # Row 17: Response and Clearing Time
    ws['A17'] = 'Response Time:'
    ws['C17'] = row.get('response_time')
    ws['E17'] = 'Clearing Time:'
    ws['F17'] = row.get('clearing_time')

    # Row 18: Department Contacted
    ws['A18'] = 'Department Contacted:'
    ws['C18'] = row.get('department_contact')

    # Make labels bold
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    thick_side = Side(style='thick')
    for row_num in range(6, 19):
        for cell in ws[row_num]:
            if cell.value and isinstance(cell.value, str) and cell.value.endswith(':'):
                cell.font = Font(bold=True)
            if cell.value:
                cell.border = thin_border

    # Add outer thick borders
    for r in range(1, 19):
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            if cell.border:
                current = cell.border
                new_border = Border(
                    left=thick_side if c == 1 else current.left,
                    right=thick_side if c == 11 else current.right,
                    top=thick_side if r == 1 else current.top,
                    bottom=thick_side if r == 18 else current.bottom
                )
                cell.border = new_border
            elif cell.value:
                cell.border = Border(
                    left=thick_side if c == 1 else Side(style='thin'),
                    right=thick_side if c == 11 else Side(style='thin'),
                    top=thick_side if r == 1 else Side(style='thin'),
                    bottom=thick_side if r == 18 else Side(style='thin')
                )

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = utils.get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

def search_page():
    import streamlit as st
    import pandas as pd
    from db_utils import get_sqlalchemy_engine
    from io import BytesIO
    import openpyxl
    from openpyxl.drawing.image import Image as OpenpyxlImage
    import os
    import tempfile # NEW: For creating temporary image files

    st.header("🔍 Search & View Data")

    search_options = ["Accidents", "Incidents", "Breaks", "Pickups"]
    selected_option = st.selectbox("Select data to view", search_options)

    start_date = st.date_input("Start date").strftime('%Y-%m-%d')
    end_date = st.date_input("End date").strftime('%Y-%m-%d')
    vehicle = st.text_input("Vehicle (optional)")

    if st.button("Search"):
        try:
            engine = get_sqlalchemy_engine()
        except Exception as e:
            st.error(f"❌ ERROR: Could not get database engine. Details: {e}")
            return # Exit if DB connection fails

        df = pd.DataFrame()
        image_df = pd.DataFrame()

        # --- 1. CONSTRUCT QUERY AND FETCH MAIN DATA ---
        if selected_option in ["Accidents", "Incidents"]:
            incident_type = "Accident" if selected_option == "Accidents" else "Incident"
            query = "SELECT * FROM incident_reports WHERE incident_date BETWEEN ? AND ? AND incident_type = ?"
            params = (start_date, end_date, incident_type)

            if vehicle:
                query += " AND patrol_car = ?"
                params = params + (vehicle,)

            if st.session_state["role"] == "re_admin":
                contractor_id = st.session_state.get("active_contractor")
                if contractor_id:
                    query += " AND contractor_id = ?"
                    params = params + (contractor_id,)

            try:
                # Fetch main report data. The 'id' column is the Primary Key.
                with engine.connect() as conn:
                    df = pd.read_sql_query(query, conn, params=params)
            except Exception as e:
                st.error(f"Database error fetching incident data: {e}")


            # --- 2. FETCH IMAGE DATA (BLOBs) FOR LINKING ---
            if not df.empty and 'id' in df.columns: # 'id' is the primary key (PK)
                report_ids = df['id'].tolist()

                # Handling the IN clause parameters
                placeholders = ','.join(['?'] * len(report_ids))

                # Query the image table for BLOB data and metadata
                # Note the change: incident_images table and incident_id column
                image_query = f"SELECT incident_id, image_name, image_data FROM incident_images WHERE incident_id IN ({placeholders})"

                try:
                    # Fetching the BLOB data is slow but necessary for embedding
                    with engine.connect() as conn:
                        image_df = pd.read_sql_query(image_query, conn, params=tuple(report_ids))
                    st.success(f"Found {len(df)} reports and {len(image_df)} associated images.")
                except Exception as e:
                    st.warning(f"Could not fetch image data from incident_images table: {e}. Download will be data-only.")

        elif selected_option == "Breaks":
            query = "SELECT * FROM breaks WHERE break_date BETWEEN ? AND ?"
            params = (start_date, end_date)
            if vehicle:
                query += " AND vehicle = ?"
                params = params + (vehicle,)
            try:
                with engine.connect() as conn:
                    df = pd.read_sql_query(query, conn, params=params)
            except Exception as e:
                st.error(f"Database error fetching breaks data: {e}")

        elif selected_option == "Pickups":
            query = "SELECT * FROM pickups WHERE DATE(pickup_start) BETWEEN ? AND ?"
            params = (start_date, end_date)
            if vehicle:
                query += " AND vehicle = ?"
                params = params + (vehicle,)
            try:
                with engine.connect() as conn:
                    df = pd.read_sql_query(query, conn, params=params)
            except Exception as e:
                st.error(f"Database error fetching pickups data: {e}")

        # --- 3. DISPLAY AND DOWNLOAD LOGIC ---
        if not df.empty:
            st.dataframe(df)

            # Excel Workbook Download
            temp_dir = None
            output = BytesIO()
            try:
                if selected_option in ["Accidents", "Incidents"] and not image_df.empty:
                    temp_dir = tempfile.mkdtemp()

                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    if selected_option in ["Accidents", "Incidents"]:
                        # Create a template sheet for each report
                        for index, row in df.iterrows():
                            contractor_id = row.get('contractor_id')
                            if contractor_id == 1:
                                rfi_prefix = 'WIZP'
                            elif contractor_id == 2:
                                rfi_prefix = 'PASC'
                            else:
                                rfi_prefix = 'PASC'
                            sheet_name = f"{rfi_prefix}{index+1:03d}"[:31]
                            ws = writer.book.create_sheet(title=sheet_name)
                            fill_incident_template(ws, row, index)
                    else:
                        # For other options, use column sheets
                        for col in df.columns:
                            sheet_name = col[:31]
                            df[[col]].to_excel(writer, sheet_name=sheet_name, index=False)

                    workbook = writer.book

                    # Embed images if applicable
                    if selected_option in ["Accidents", "Incidents"] and not image_df.empty and temp_dir:
                        for index, row in df.iterrows():
                            report_id = row.get('id')
                            if report_id:
                                contractor_id = row.get('contractor_id')
                                if contractor_id == 1:
                                    rfi_prefix = 'WIZP'
                                elif contractor_id == 2:
                                    rfi_prefix = 'PASC'
                                else:
                                    rfi_prefix = 'PASC'
                                sheet_name = f"{rfi_prefix}{index+1:03d}"[:31]
                                ws = writer.sheets[sheet_name]

                                linked_images = image_df[image_df['incident_id'] == report_id]

                                if not linked_images.empty:
                                    current_row = 19  # Below the template
                                    ws.cell(row=current_row, column=1, value=f"--- IMAGES FOR REPORT ID: {report_id} ---").font = openpyxl.styles.Font(bold=True)
                                    current_row += 1

                                    for img_index, img_row in linked_images.iterrows():
                                        img_data_blob = img_row['image_data']
                                        img_name = img_row['image_name']

                                        if not img_data_blob or not isinstance(img_data_blob, bytes) or len(img_data_blob) == 0:
                                            continue

                                        file_ext = img_name.split('.')[-1] if '.' in img_name else 'png'

                                        temp_file_path = os.path.join(temp_dir, f"{report_id}_{img_index}.{file_ext}")
                                        try:
                                            with open(temp_file_path, 'wb') as f:
                                                f.write(img_data_blob)

                                            img = OpenpyxlImage(temp_file_path)
                                            img.anchor = f'A{current_row}'
                                            img.width = 400
                                            img.height = 300
                                            ws.add_image(img)

                                            current_row += 10
                                        except Exception as img_e:
                                            # Skip this image and continue
                                            continue

                                    current_row += 2

                # --- DOWNLOAD BUTTON FOR EXCEL WORKBOOK ---
                st.download_button(
                    label="⬇️ Download Excel Workbook",
                    data=output.getvalue(),
                    file_name=f"{selected_option.lower()}_results_workbook.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            except Exception as e:
                st.error(f"Error creating Excel workbook: {e}")
                st.warning("Ensure all data is valid and required libraries are installed.")

            finally:
                if temp_dir:
                    import shutil
                    shutil.rmtree(temp_dir)

            # CSV Download
            csv = df.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="⬇️ Download Data as CSV",
                data=csv,
                file_name=f"{selected_option.lower()}_results.csv",
                mime="text/csv"
            )
        else:
            st.info("No data found for your search.")
